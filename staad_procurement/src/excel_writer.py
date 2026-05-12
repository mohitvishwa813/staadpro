"""
Excel Procurement Report Generator
Creates professional Excel output from parsed STAAD data
"""

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import os


# ── Colors ────────────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F4E79"   # dark blue
CLR_HEADER_FG   = "FFFFFF"
CLR_SUBHDR_BG   = "2E75B6"
CLR_SUBHDR_FG   = "FFFFFF"
CLR_GROUP_BG    = "D6E4F0"   # light blue — thickness group rows
CLR_ALT_ROW     = "EBF3FB"
CLR_TOTAL_BG    = "FCE4D6"   # light orange for totals
CLR_TITLE_BG    = "0D2137"
CLR_ACCENT      = "ED7D31"

THIN = Side(style='thin', color="BFBFBF")
MED  = Side(style='medium', color="2E75B6")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def _hdr_font(size=11, bold=True, color=CLR_HEADER_FG):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _cell_font(size=10, bold=False, color="000000"):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def _right():
    return Alignment(horizontal='right', vertical='center')

def _left():
    return Alignment(horizontal='left', vertical='center')

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Sheet 1: Member-wise Detail ────────────────────────────────────────────────
def write_detail_sheet(ws, records, filename):
    ws.title = "Member Detail"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"STAAD.Pro — Material Procurement Detail"
    ws["A1"].font   = Font(name='Arial', size=14, bold=True, color=CLR_HEADER_FG)
    ws["A1"].fill   = _fill(CLR_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = f"File: {filename}    |    Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font      = Font(name='Arial', size=9, italic=True, color="AAAAAA")
    ws["A2"].fill      = _fill(CLR_TITLE_BG)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 16

    # Column headers
    headers = ["Member ID", "Part", "Thickness\n(mm)", "Width\n(m)",
                "Length\n(m)", "Area\n(m²)", "Weight\n(kg)", "Grade"]
    ws.row_dimensions[3].height = 34
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = _hdr_font(10)
        c.fill      = _fill(CLR_SUBHDR_BG)
        c.alignment = _center(wrap=True)
        c.border    = BORDER_MED

    # Data rows
    for ri, rec in enumerate(records, 4):
        alt = (ri % 2 == 0)
        bg  = CLR_ALT_ROW if alt else "FFFFFF"
        row_vals = [
            rec['Member ID'],
            rec['Part'],
            rec['Thickness (mm)'],
            rec['Width (m)'],
            rec['Length (m)'],
            rec['Area (m²)'],
            rec['Weight (kg)'],
            "IS 2062 E250",
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = _cell_font()
            c.fill   = _fill(bg)
            c.border = BORDER_THIN
            if ci == 1:
                c.alignment = _center()
            elif ci in (3, 4, 5, 6, 7):
                c.alignment = _right()
                if ci == 7:
                    c.number_format = '#,##0.00'
                elif ci in (6,):
                    c.number_format = '0.0000'
            else:
                c.alignment = _left()

    set_col_widths(ws, {
        'A': 12, 'B': 16, 'C': 13, 'D': 10,
        'E': 10, 'F': 11, 'G': 12, 'H': 14
    })

    ws.freeze_panes = "A4"


# ── Sheet 2: Thickness-wise Summary ───────────────────────────────────────────
def write_summary_sheet(ws, records):
    ws.title = "Procurement Summary"
    ws.sheet_view.showGridLines = False

    # Group by thickness
    by_thk = defaultdict(lambda: {'count': 0, 'total_area': 0.0, 'total_weight': 0.0, 'members': set()})
    for rec in records:
        thk = rec['Thickness (mm)']
        by_thk[thk]['count']        += 1
        by_thk[thk]['total_area']   += rec['Area (m²)']
        by_thk[thk]['total_weight'] += rec['Weight (kg)']
        by_thk[thk]['members'].add(rec['Member ID'])

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "STAAD.Pro — Thickness-wise Procurement Summary"
    ws["A1"].font      = Font(name='Arial', size=14, bold=True, color=CLR_HEADER_FG)
    ws["A1"].fill      = _fill(CLR_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = "GROUP BY PLATE THICKNESS"
    ws["A2"].font      = Font(name='Arial', size=9, bold=True, color=CLR_ACCENT)
    ws["A2"].fill      = _fill("0D2137")
    ws["A2"].alignment = _center()

    headers = ["Thickness\n(mm)", "No. of\nPlate Pieces",
               "No. of\nMembers", "Total Area\n(m²)",
               "Wt per Plate\n(kg)", "Total Weight\n(kg)", "Total Weight\n(Ton)", "Grade"]
    ws.row_dimensions[3].height = 36
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = _hdr_font(10)
        c.fill      = _fill(CLR_SUBHDR_BG)
        c.alignment = _center(wrap=True)
        c.border    = BORDER_MED

    total_w = 0.0
    total_a = 0.0
    for ri, (thk, data) in enumerate(sorted(by_thk.items()), 4):
        bg = CLR_GROUP_BG if ri % 2 == 0 else "FFFFFF"
        count = data['count']
        wt_per_plate = round(data['total_weight'] / count, 2) if count else 0.0
        row_vals = [
            f"{thk} mm",
            count,
            len(data['members']),
            round(data['total_area'], 3),
            wt_per_plate,
            round(data['total_weight'], 2),
            round(data['total_weight'] / 1000, 3),
            "IS 2062 E250",
        ]
        total_w += data['total_weight']
        total_a += data['total_area']
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = Font(name='Arial', size=11, bold=(ci == 1))
            c.fill   = _fill(bg)
            c.border = BORDER_THIN
            c.alignment = _center() if ci in (1, 2, 3, 8) else _right()
            if ci in (5, 6):
                c.number_format = '#,##0.00'
            elif ci == 7:
                c.number_format = '0.000'

    # Grand total row
    tr = 4 + len(by_thk)
    grand = ["GRAND TOTAL", "", "", round(total_a, 3),
             "", round(total_w, 2), round(total_w/1000, 3), ""]
    for ci, v in enumerate(grand, 1):
        c = ws.cell(row=tr, column=ci, value=v)
        c.font   = Font(name='Arial', size=11, bold=True, color="000000")
        c.fill   = _fill(CLR_TOTAL_BG)
        c.border = BORDER_MED
        c.alignment = _center() if ci in (1, 2, 3, 8) else _right()
        if ci == 6:
            c.number_format = '#,##0.00'
        elif ci == 7:
            c.number_format = '0.000'

    set_col_widths(ws, {
        'A': 14, 'B': 16, 'C': 14,
        'D': 14, 'E': 18, 'F': 16, 'G': 16, 'H': 14
    })
    ws.freeze_panes = "A4"


# ── Sheet 3: Part-wise Breakup ─────────────────────────────────────────────────
def write_partwise_sheet(ws, records):
    ws.title = "Part-wise Breakup"
    ws.sheet_view.showGridLines = False

    by_part = defaultdict(lambda: {'weight': 0.0, 'area': 0.0, 'count': 0})
    for rec in records:
        key = (rec['Part'], rec['Thickness (mm)'])
        by_part[key]['weight'] += rec['Weight (kg)']
        by_part[key]['area']   += rec['Area (m²)']
        by_part[key]['count']  += 1

    ws.merge_cells("A1:F1")
    ws["A1"] = "Part-wise + Thickness-wise Breakup"
    ws["A1"].font      = Font(name='Arial', size=13, bold=True, color=CLR_HEADER_FG)
    ws["A1"].fill      = _fill(CLR_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Part", "Thickness (mm)", "Count", "Area (m²)", "Weight (kg)", "Weight (Ton)"]
    ws.row_dimensions[2].height = 24
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _hdr_font(10)
        c.fill      = _fill(CLR_SUBHDR_BG)
        c.alignment = _center(wrap=True)
        c.border    = BORDER_MED

    for ri, ((part, thk), data) in enumerate(sorted(by_part.items()), 3):
        alt = ri % 2 == 0
        bg  = CLR_ALT_ROW if alt else "FFFFFF"
        row = [part, f"{thk} mm", data['count'],
               round(data['area'], 3), round(data['weight'], 2),
               round(data['weight']/1000, 3)]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = _cell_font()
            c.fill   = _fill(bg)
            c.border = BORDER_THIN
            c.alignment = _right() if ci > 2 else _left()
            if ci == 5:
                c.number_format = '#,##0.00'

    set_col_widths(ws, {'A': 18, 'B': 16, 'C': 10, 'D': 14, 'E': 14, 'F': 14})
    ws.freeze_panes = "A3"


# ── Sheet 4: Pipes ─────────────────────────────────────────────────────────────
def write_pipes_sheet(ws, pipe_records):
    ws.title = "Pipes"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "STAAD.Pro — Pipe / Hollow Round Sections"
    ws["A1"].font      = Font(name='Arial', size=14, bold=True, color=CLR_HEADER_FG)
    ws["A1"].fill      = _fill(CLR_TITLE_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    headers = ["Member ID", "OD (mm)", "Wall Thickness (mm)",
               "Length (m)", "Surface Area (m²)", "Weight (kg)"]
    ws.row_dimensions[2].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _hdr_font(10)
        c.fill      = _fill(CLR_SUBHDR_BG)
        c.alignment = _center(wrap=True)
        c.border    = BORDER_MED

    total_wt  = 0.0
    total_len = 0.0
    for ri, rec in enumerate(pipe_records, 3):
        alt = ri % 2 == 0
        bg  = CLR_ALT_ROW if alt else "FFFFFF"
        row = [
            rec['Member ID'],
            rec['OD (mm)'],
            rec['Wall Thickness (mm)'],
            rec['Length (m)'],
            rec['Surface Area (m²)'],
            rec['Weight (kg)'],
        ]
        total_wt  += rec['Weight (kg)']
        total_len += rec['Length (m)']
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = _cell_font()
            c.fill   = _fill(bg)
            c.border = BORDER_THIN
            c.alignment = _center() if ci == 1 else _right()
            if ci == 6:
                c.number_format = '#,##0.00'
            elif ci in (4, 5):
                c.number_format = '0.000'

    if pipe_records:
        tr = 3 + len(pipe_records)
        grand = ["GRAND TOTAL", "", "", round(total_len, 3), "", round(total_wt, 2)]
        for ci, v in enumerate(grand, 1):
            c = ws.cell(row=tr, column=ci, value=v)
            c.font   = Font(name='Arial', size=11, bold=True)
            c.fill   = _fill(CLR_TOTAL_BG)
            c.border = BORDER_MED
            c.alignment = _center() if ci == 1 else _right()
            if ci == 6:
                c.number_format = '#,##0.00'
            elif ci == 4:
                c.number_format = '0.000'

    set_col_widths(ws, {'A': 12, 'B': 12, 'C': 20, 'D': 12, 'E': 18, 'F': 14})
    ws.freeze_panes = "A3"


# ── Master function ─────────────────────────────────────────────────────────────
def generate_excel(records, filename, output_path, pipe_records=None):
    wb = Workbook()
    ws1 = wb.active
    write_detail_sheet(ws1, records, filename)

    ws2 = wb.create_sheet()
    write_summary_sheet(ws2, records)

    ws3 = wb.create_sheet()
    write_partwise_sheet(ws3, records)

    if pipe_records:
        ws4 = wb.create_sheet()
        write_pipes_sheet(ws4, pipe_records)

    wb.save(output_path)
    return output_path
